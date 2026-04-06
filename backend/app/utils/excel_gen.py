from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Any

NAVY_HEX = "1E3A5F"
LIGHT_BLUE_HEX = "DBE8F5"
WHITE_HEX = "FFFFFF"

navy_fill = PatternFill("solid", fgColor=NAVY_HEX)
blue_fill = PatternFill("solid", fgColor=LIGHT_BLUE_HEX)
white_fill = PatternFill("solid", fgColor=WHITE_HEX)

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _val(v):
    return str(v) if v is not None and v != "" else ""


def _bool_label(v):
    if v is True:
        return "Yes"
    if v is False:
        return "No"
    return ""


def _section(ws, row, title):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, color=WHITE_HEX, size=11)
    cell.fill = navy_fill
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 20
    return row + 1


def _row(ws, row, label, value, label_col=1, value_col=2, merge_value=False):
    lc = ws.cell(row=row, column=label_col, value=label)
    lc.font = Font(bold=True, size=9)
    lc.fill = blue_fill
    lc.border = thin_border
    lc.alignment = Alignment(wrap_text=True)

    vc = ws.cell(row=row, column=value_col, value=value)
    vc.font = Font(size=9)
    vc.fill = white_fill
    vc.border = thin_border
    vc.alignment = Alignment(wrap_text=True)

    if merge_value:
        ws.merge_cells(start_row=row, start_column=value_col, end_row=row, end_column=4)

    return row + 1


def generate_excel(submission: Any) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tax Questionnaire"

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22

    # Title
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "Machin & Co. Tax Services"
    title_cell.font = Font(bold=True, size=16, color=NAVY_HEX)
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    sub_cell = ws["A2"]
    sub_cell.value = "Personal Income Tax Questionnaire"
    sub_cell.font = Font(bold=True, size=12, color=NAVY_HEX)
    sub_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 22

    r = 4

    # ─── Personal Info ───────────────────────────────────────────────────────
    r = _section(ws, r, "Personal Information")

    pairs = [
        ("Taxpayer's Name", _val(submission.taxpayer_name), "Social Security", _val(submission.taxpayer_ssn)),
        ("Spouse Name", _val(submission.spouse_name), "Spouse SSN", _val(submission.spouse_ssn)),
        ("Address", _val(submission.address), "City", _val(submission.city)),
        ("State", _val(submission.state), "Zip", _val(submission.zip_code)),
        ("Telephone", _val(submission.telephone), "Work Phone", _val(submission.work_phone)),
        ("Date of Birth", _val(submission.date_of_birth), "Occupation", _val(submission.occupation)),
        ("Spouse DOB", _val(submission.spouse_dob), "Spouse Occupation", _val(submission.spouse_occupation)),
        ("E-Mail", _val(submission.email), "Filing Status", _val(submission.filing_status)),
        ("How did you hear about us?", _val(submission.how_heard), "", ""),
    ]
    for lbl1, val1, lbl2, val2 in pairs:
        _row(ws, r, lbl1, val1, label_col=1, value_col=2)
        _row(ws, r, lbl2, val2, label_col=3, value_col=4)
        r += 1

    r += 1

    # ─── Dependents ──────────────────────────────────────────────────────────
    r = _section(ws, r, "Dependents")
    headers = ["Name", "Relationship", "SSN / DOB", "Months at Home"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = Font(bold=True, size=9)
        c.fill = blue_fill
        c.border = thin_border
    r += 1
    for d in (submission.dependents or []):
        vals = [_val(d.name), _val(d.relationship), f"{_val(d.ssn)} / {_val(d.date_of_birth)}", _val(d.months_lived_home)]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.font = Font(size=9)
            c.border = thin_border
        r += 1
    r += 1

    # ─── Income ──────────────────────────────────────────────────────────────
    r = _section(ws, r, "Income (checked items)")
    income_keys = [
        ("wage_w2", "Wage Statements W-2s"), ("f1099s", "1099s"), ("iras", "IRAs"),
        ("sale_investments", "Sale of Investments"), ("unemployment", "Received Unemployment"), ("alimony", "Alimony"),
        ("buy_sell_home", "Buy or Sell of a Home"), ("rental_property", "Own Rental Property"), ("interest_received", "Interest Received"),
        ("dividends", "Dividends Received"), ("pension", "Pension or Retirement Income"), ("social_security", "Social Security Income"),
        ("self_employed", "Self Employed/Own a Business"), ("tips_other", "Tips/Other Income"), ("farm_income", "Farm Income"),
        ("household_income", "Household Income"), ("lottery", "Lottery or Gambling Winnings"), ("corporate", "Corporate Income"),
    ]
    income_list = submission.income_types or []
    for k, label in income_keys:
        checked = "Yes" if k in income_list else "No"
        r = _row(ws, r, label, checked, label_col=1, value_col=2)
    r += 1

    # ─── Deductions ──────────────────────────────────────────────────────────
    r = _section(ws, r, "Deductions (checked items)")
    deduction_keys = [
        ("property_tax", "Property Tax"), ("union_dues", "Union Dues"), ("moving_expense", "Moving Expense"),
        ("medical", "Medical Expense"), ("job_related", "Job Related Expense"), ("education", "Education Expense"),
        ("mortgage_interest", "Mortgage Interest"), ("significant_loss", "Significant Loss or Theft"), ("tax_prep", "Tax Preparation Expense"),
        ("charity", "Charity or Religious Contribution"), ("retirement_savings", "Retirement Savings"), ("mortgage_points", "Mortgage Points"),
    ]
    ded_list = submission.deduction_types or []
    for k, label in deduction_keys:
        checked = "Yes" if k in ded_list else "No"
        r = _row(ws, r, label, checked, label_col=1, value_col=2)
    r += 1

    # ─── Child Tax Credit ────────────────────────────────────────────────────
    r = _section(ws, r, "Child Tax Credit")
    ctc_fields = [
        ("Claiming child for Child Tax Credit/EIC?", _bool_label(submission.claiming_child_credit)),
        ("Biological parent?", _bool_label(submission.biological_parent)),
        ("Has custody?", _bool_label(submission.has_custody)),
        ("Other parent location", _val(submission.other_parent_location)),
        ("Why parents not claiming?", _val(submission.why_parents_not_claiming)),
        ("Biological parents location", _val(submission.biological_parents_location)),
        ("Other parent works?", _val(submission.other_parent_works)),
        ("Other parent income", _val(submission.other_parent_income)),
        ("Benefits received", ", ".join(submission.benefits_received or [])),
        ("Caretaker while working", _val(submission.caretaker_while_working)),
    ]
    for label, value in ctc_fields:
        r = _row(ws, r, label, value, merge_value=True)
    r += 1

    # ─── Child Care ──────────────────────────────────────────────────────────
    r = _section(ws, r, "Child Care Information")
    for label, value in [
        ("Provider's Name", _val(submission.provider_name)),
        ("Provider's SSN/EIN", _val(submission.provider_ssn_ein)),
        ("Provider's Address", _val(submission.provider_address)),
        ("Provider Telephone", _val(submission.provider_phone)),
        ("City, State, Zip", _val(submission.provider_city_state_zip)),
        ("Amount Paid", _val(submission.amount_paid_provider)),
    ]:
        r = _row(ws, r, label, value, merge_value=True)
    r += 1

    # ─── Educational Expenses ────────────────────────────────────────────────
    r = _section(ws, r, "Educational Expenses")
    for label, value in [
        ("School Name", _val(submission.school_name)),
        ("School Phone", _val(submission.school_phone)),
        ("School Address", _val(submission.school_address)),
        ("City, State, Zip", _val(submission.school_city_state_zip)),
        ("Tuition Paid", _val(submission.tuition_paid)),
        ("Supplies Bought", _val(submission.supplies_bought)),
        ("Equipment Bought", _val(submission.equipment_bought)),
        ("Person Attending", _val(submission.person_attending)),
    ]:
        r = _row(ws, r, label, value, merge_value=True)
    r += 1

    # ─── Supplementary ───────────────────────────────────────────────────────
    r = _section(ws, r, "Supplementary Statements")
    r = _row(ws, r, "Delinquent on federal loans?", _bool_label(submission.delinquent_loans), merge_value=True)
    r = _row(ws, r, "IRS Debt?", _bool_label(submission.irs_debt), merge_value=True)
    r += 1

    # ─── Other Income ────────────────────────────────────────────────────────
    r = _section(ws, r, "Other Sources of Income")
    other_income_keys = [
        ("self_employed", "Self-Employed"), ("llc", "LLC"), ("f1120", "1120"),
        ("f1120s", "1120S"), ("f1065", "1065"), ("f1099_recipient", "1099 Recipient"),
    ]
    other_list = submission.other_income_types or []
    for k, label in other_income_keys:
        r = _row(ws, r, label, "Yes" if k in other_list else "No")
    r += 1

    # ─── Business Income & Expense ───────────────────────────────────────────
    r = _section(ws, r, "Business Income & Expense")
    biz_fields = [
        ("Business/Profession", _val(submission.business_name)),
        ("Business Address", _val(submission.business_address)),
        ("City, State, Zip", _val(submission.business_city_state_zip)),
        ("Employer ID", _val(submission.employer_id)),
        ("Business Owned By", _val(submission.business_owned_by)),
        ("Accounting Method", _val(submission.accounting_method)),
        ("Total Gross Receipts", _val(submission.gross_receipts)),
        ("Returns and Allowance", _val(submission.returns_allowance)),
        ("Cost of Goods Sold", _val(submission.cost_goods_sold)),
        ("Gross Profit", _val(submission.gross_profit)),
        ("Advertising", _val(submission.expense_advertising)),
        ("Car and Truck Expenses", _val(submission.expense_car_truck)),
        ("Commissions and Fees", _val(submission.expense_commissions)),
        ("Contract Labor", _val(submission.expense_contract_labor)),
        ("Depreciation", _val(submission.expense_depreciation)),
        ("Insurance", _val(submission.expense_insurance)),
        ("Legal and Professional", _val(submission.expense_legal)),
        ("Office Expenses", _val(submission.expense_office)),
        ("Rent or Lease", _val(submission.expense_rent_lease)),
        ("Repair/Maintenance", _val(submission.expense_repair)),
        ("Supplies", _val(submission.expense_supplies)),
        ("Taxes and License Fee", _val(submission.expense_taxes_license)),
        ("Travel", _val(submission.expense_travel)),
        ("Meals and Entertainment", _val(submission.expense_meals)),
        ("Utilities", _val(submission.expense_utilities)),
        ("Wages", _val(submission.expense_wages)),
        ("Other Expenses", _val(submission.expense_other)),
    ]
    for label, value in biz_fields:
        r = _row(ws, r, label, value, merge_value=True)
    r += 1

    # ─── Banking ─────────────────────────────────────────────────────────────
    r = _section(ws, r, "Banking Information")
    for label, value in [
        ("Bank Name", _val(submission.bank_name)),
        ("Routing Number", _val(submission.routing_number)),
        ("Account Number", _val(submission.account_number)),
    ]:
        r = _row(ws, r, label, value, merge_value=True)
    r += 1

    # ─── Signatures ──────────────────────────────────────────────────────────
    r = _section(ws, r, "Signatures")
    for label, value in [
        ("Taxpayer's Signature", _val(submission.taxpayer_signature)),
        ("Taxpayer Signature Date", _val(submission.taxpayer_signature_date)),
        ("Spouse Signature", _val(submission.spouse_signature)),
        ("Spouse Signature Date", _val(submission.spouse_signature_date)),
    ]:
        r = _row(ws, r, label, value, merge_value=True)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
